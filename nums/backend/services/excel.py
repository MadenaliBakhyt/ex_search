"""Excel helpers: preview files, read company names, write results."""
from __future__ import annotations

import io
import logging
from typing import Iterable

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

MAX_COMPANIES = 10000  # cap to avoid runaway processing
PREVIEW_ROWS = 5     # number of rows returned by preview_excel


def _load(data: bytes):
    try:
        return load_workbook(
            filename=io.BytesIO(data), data_only=True, read_only=True
        )
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}") from e


def _col_letter(idx: int) -> str:
    """0 -> A, 1 -> B, 25 -> Z, 26 -> AA, ..."""
    result = ""
    n = idx
    while True:
        result = chr(ord("A") + (n % 26)) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


def preview_excel(data: bytes, max_rows: int = PREVIEW_ROWS) -> dict:
    """Return a lightweight preview: column letters and the first few rows.

    Shape:
        {
          "columns": ["A", "B", "C", ...],
          "rows":    [["foo", "bar", ...], ...],
          "total_columns": N,
        }
    """
    wb = _load(data)
    ws = wb.active
    if ws is None:
        return {"columns": [], "rows": [], "total_columns": 0}

    rows: list[list[str]] = []
    num_cols = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = ["" if v is None else str(v) for v in row]
        # Trim trailing empty cells so we don't report phantom columns
        while cells and cells[-1] == "":
            cells.pop()
        rows.append(cells)
        num_cols = max(num_cols, len(cells))

    # Normalize row lengths so the frontend can render a rectangular table
    for row in rows:
        while len(row) < num_cols:
            row.append("")

    columns = [_col_letter(i) for i in range(num_cols)]
    return {"columns": columns, "rows": rows, "total_columns": num_cols}


def read_company_names(
    data: bytes,
    *,
    column_index: int = 0,
    skip_first_row: bool = False,
    dedupe: bool = True,
) -> list[str]:
    """Return company names from the chosen column of the uploaded xlsx.

    Parameters
    ----------
    column_index:
        Zero-based index of the column to read.
    skip_first_row:
        If True, skip the first row (assumed to be a header like "Company").
    dedupe:
        If True, drop case-insensitive duplicate names.
    """
    wb = _load(data)
    ws = wb.active
    if ws is None:
        return []

    if column_index < 0:
        raise ValueError("column_index must be >= 0")

    col = column_index + 1  # openpyxl is 1-indexed
    min_row = 2 if skip_first_row else 1

    seen: set[str] = set()
    names: list[str] = []

    for row in ws.iter_rows(
        min_row=min_row, min_col=col, max_col=col, values_only=True
    ):
        if not row:
            continue
        value = row[0]
        if value is None:
            continue

        name = str(value).strip()
        if not name:
            continue

        if dedupe:
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)

        names.append(name)

        if len(names) >= MAX_COMPANIES:
            logger.info("hit MAX_COMPANIES=%d, truncating input", MAX_COMPANIES)
            break

    return names


def build_results_workbook(rows: Iterable[dict]) -> bytes:
    """Serialize result rows to an xlsx byte blob."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Company", "Website", "Phone"])

    for r in rows:
        ws.append([
            r.get("company") or "",
            r.get("website") or "",
            r.get("phone") or "",
        ])

    widths = {"A": 40, "B": 45, "C": 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
